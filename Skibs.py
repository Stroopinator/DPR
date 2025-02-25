import statbotics 
import csv
import pandas as pd
from openpyxl import load_workbook
from IPython.display import display
from tabulate import tabulate
import numpy as np
import sys
from tabulate import tabulate


class team:
    def __init__(self, number):
        self._matchDPR = []
        self._number = number
        self._totalDPR = 0

    def getNumber(self):
        return self._number         
    
    def addMatch(self, DPR):
        self._matchDPR.append(DPR) 
        self._totalDPR += DPR

    def getSDV(self):
        return round(np.std(a = self._matchDPR, ddof = 1),2)
         

def getMatchDPR(match, color):

    sum = 0

    if color:
        for t in match['alliances']['blue']['team_keys']:
            sum += teams[t]['epa']['breakdown']['teleop_points'] 
        return sum - match['result']["blue_teleop_points"]
    else:
        for t in match['alliances']['red']['team_keys']:
            sum += teams[t]['epa']['breakdown']['teleop_points'] 
        return sum - match['result']["red_teleop_points"]





file_path = 'Stats/DataFrame.xlsx'
workBook = load_workbook(file_path)
print(workBook.sheetnames)

# Delete all sheets
for sheet_name in workBook.sheetnames:
    if sheet_name != "Sheet1" : del workBook[sheet_name]

workBook.save(file_path)
print(workBook.sheetnames)



# Save the workbook
sb = statbotics.Statbotics()



while(True):
    _event = input("Enter Comp: ")
    if _event == "quit": break

    print(_event)

    try:
        matches = sb.get_matches(event = _event, limit = 1000)
        Tempteams = sb.get_team_events(event = _event, limit = 1000)
        eventName = sb.get_event(event=_event)
    except:
        print("invalid key")
        continue

    if _event in workBook.sheetnames:
        print("Competition already ran")
        continue

    print(workBook.sheetnames)


    teams = {}

    for t in Tempteams:
        teams[t['team']] = t
  

    teamDict = {}
    tempDict = {'Number': [],
            'DPR': [],
            'EPA' : [],
            'SDV' : []}
    SDVDict = {}
    dataOut = []


    matchArray = []
    matchDPRArray = []

    tempTeamNum = 0

    #this is changed from t["team"] to just t for the workaround
    #TODO: change this back when the API gets fixed
    for t in Tempteams:
        cur = team(t['team'])
        SDVDict[t['team']] = cur
        teamDict[t['team']] = tempTeamNum
        tempTeamNum += 1







    #print(sb.get_team_events(team = 5827, year = 2024))
        #could have broken something here
    matchCount = 0
    for m in matches:
        if m['result']['red_teleop_points'] != None and m['result']['blue_teleop_points'] != None and m["comp_level"] == 'qm':
            tempTeams = []
            if (getMatchDPR(m, True) <= eventName["epa"]["top_8"]):
                for t in m['alliances']['red']['team_keys']:
                    tempTeams.append(t)
                    SDVDict[t].addMatch(getMatchDPR(m, True))

                matchArray.append(tempTeams)
                matchDPRArray.append(getMatchDPR(m, True))
                matchCount += 1

            tempTeams = []
            if getMatchDPR(m, False) <= eventName["epa"]["top_8"]:
                for t in m['alliances']['blue']['team_keys']:
                    tempTeams.append(t)
                    SDVDict[t].addMatch(getMatchDPR(m, False))

                    #not sure whether this line goes in the loop or not
                matchArray.append(tempTeams)
                matchDPRArray.append(getMatchDPR(m, False))
                matchCount += 1

    #DO NOT TOCUH PLEASE
        
    #print(matchArray)
    #print(matchDPRArray)
            
    NDMatchArray = np.zeros((len(matchArray), len(teams)))
    NDMatchDPRArray = np.zeros(len(matchArray))
    print("pruned " + str(2*len(matches) - (len(matchArray))))

    for i in range(len(matchArray)):
        NDMatchDPRArray[i] = matchDPRArray[i]
        NDMatchArray[i][teamDict[matchArray[i][0]]] = 1
        NDMatchArray[i][teamDict[matchArray[i][1]]] = 1
        NDMatchArray[i][teamDict[matchArray[i][2]]] = 1 

    dprs = np.linalg.lstsq(NDMatchArray, NDMatchDPRArray, rcond=None)   

    floor = sys.maxsize
    ceiling = -sys.maxsize - 1
    for dpr in dprs[0]:
        ceiling = max(ceiling, dpr)
        floor = min(floor, dpr)




    #matchNorm = np.dot(np.transpose(matchArray), matchArray)
    #DPRNorm = np.dot(np.transpose(matchArray), matchDPRArray)
        
    #like all other changes this was made to workaround get_teams_events being broken chagne back when it is fixed
    #TODO: fix this later
    for t in teams:
        #dataOut({t["team"], dprs[0][teamDict[t["team"]]]})
        tempDict['EPA'].append(teams[t]['epa']['breakdown']['total_points'])
        tempDict["Number"].append(t)
        tempDict["DPR"].append(round((dprs[0][teamDict[t]] + abs(floor)) / (abs(floor) + ceiling),2))
        tempDict["SDV"].append(SDVDict[t].getSDV())
        #tempDict["Name"].append(name_dict[t["team"]])
        


        
    #print(np.linalg.solve(matchNorm, DPRNorm))


    print(dprs[1])
    df = pd.DataFrame(tempDict)
    sorted_df = df.sort_values(by='DPR', ascending=False)
    pd.set_option('display.max_rows', None)
    print(tabulate(sorted_df, headers = 'keys', tablefmt = 'psql', showindex=False))

    #new_sheet = workBook.create_sheet(title=_event)

    #workBook.save(file_path)

    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        workBook.create_sheet(title=_event)
        sorted_df.to_excel(writer, index=False, sheet_name=_event)

    print(workBook.sheetnames)




